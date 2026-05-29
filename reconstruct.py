# -*- coding: utf-8 -*-
"""Восстановление дневной стоимости портфеля из журнала 'Факт' x исторических цен MOEX.
Доходность считается методом TWR (взвешенная по времени) по дневным значениям —
ошибки/пополнения распределяются корректно. Сравнение с индексом полной доходности MCFTR."""
import openpyxl, os, json, datetime, re, urllib.request, ssl, collections, bisect, time

folder=os.path.dirname(os.path.abspath(__file__))
try:
    urllib.request.urlopen("https://iss.moex.com",timeout=10); CTX=ssl.create_default_context()
except Exception: CTX=ssl._create_unverified_context()
def get_json(url):
    for _ in range(3):
        try:
            req=urllib.request.Request(url,headers={"User-Agent":"Mozilla/5.0 dash"})
            with urllib.request.urlopen(req,timeout=40,context=CTX) as r: return json.loads(r.read().decode("utf-8"))
        except Exception as e:
            time.sleep(1); last=e
    raise last
def hist(engine,market,secid,cols,board=None):
    """вернуть [(date, {col:val})] по истории; пагинация"""
    base=f"https://iss.moex.com/iss/history/engines/{engine}/markets/{market}"
    base+=f"/boards/{board}" if board else ""
    base+=f"/securities/{secid}.json?from=2023-12-25&iss.meta=off&history.columns=TRADEDATE,{cols}"
    out=[]; start=0
    while True:
        j=get_json(base+f"&start={start}")
        b=j["history"]; rows=b["data"]; cset=b["columns"]
        if not rows: break
        for row in rows:
            d=dict(zip(cset,row)); out.append((d["TRADEDATE"],d))
        start+=len(rows)
        if len(rows)<100: break
    return out
def filler(pairs):
    pairs=sorted((d,v) for d,v in pairs if v is not None)
    ds=[d for d,_ in pairs]; vs=[v for _,v in pairs]
    def f(ds_):
        i=bisect.bisect_right(ds,ds_)-1
        return vs[i] if i>=0 else None
    return f

def adjust_splits(pairs):
    """привести исторические цены акции к текущему числу акций (учёт сплитов).
    Сплит = резкое падение цены в ~10/100/1000 раз, после которого цена осталась внизу."""
    pairs=sorted((d,v) for d,v in pairs if v is not None)
    px=[v for _,v in pairs]; splits=[]
    for i in range(1,len(px)):
        if px[i] and px[i-1] and px[i]>0:
            r=px[i-1]/px[i]
            if r>4:
                for f in (10,100,1000):
                    if abs(r-f)/f<0.08:
                        look=px[i+1:i+6]
                        if look and (sum(look)/len(look))/px[i] < 3:   # цена осталась внизу
                            splits.append((i,f)); break
    if not splits: return pairs
    out=[]
    for i,(d,p) in enumerate(pairs):
        Fa=1
        for si,f in splits:
            if si>i: Fa*=f
        out.append((d,p/Fa))
    return out

# ---------- скачать актуальные таблицы из Google (сервер за границей имеет доступ) ----------
SHEETS={
 "Antifragile.xlsx": os.environ.get("SHEET_ANTIFRAGILE",""),
 "журнал долгосрочных сделок.xlsx": os.environ.get("SHEET_ZHURNAL",""),
}
if os.environ.get("FETCH_SHEETS","1")!="0":
    for _fn,_sid in SHEETS.items():
        if not _sid: continue   # ID берём из окружения; нет ID -> используем локальную копию
        try:
            _req=urllib.request.Request(f"https://docs.google.com/spreadsheets/d/{_sid}/export?format=xlsx",
                                        headers={"User-Agent":"Mozilla/5.0"})
            with urllib.request.urlopen(_req,timeout=60,context=CTX) as _r: _data=_r.read()
            if _data[:2]==b"PK":
                open(os.path.join(folder,_fn),"wb").write(_data); print(f"скачано из Google: {_fn} ({len(_data)} б)")
            else:
                print(f"ВНИМАНИЕ: {_fn} не xlsx (нет доступа по ссылке?) — беру локальную копию")
        except Exception as _e:
            print(f"ВНИМАНИЕ: не скачал {_fn}: {_e} — беру локальную копию")

# ---------- движения и потоки из Факта ----------
av=openpyxl.load_workbook(os.path.join(folder,"Antifragile.xlsx"),data_only=True); fakt=av["Факт"]
moves=collections.defaultdict(list)   # (type,ticker) -> [(date_str, signed_qty)]
flows=collections.defaultdict(float)  # date_str -> сумма пополнения(+)/вывода(-), без первоначальной загрузки
def ds(dt): return dt.strftime("%Y-%m-%d") if hasattr(dt,'strftime') else str(dt)
for r in range(3,fakt.max_row+1):
    dt=fakt.cell(row=r,column=1).value; typ=fakt.cell(row=r,column=2).value; tic=fakt.cell(row=r,column=3).value
    mv=fakt.cell(row=r,column=4).value; qty=fakt.cell(row=r,column=5).value; g=fakt.cell(row=r,column=7).value
    if not hasattr(dt,'strftime'): continue
    d=ds(dt)
    if isinstance(tic,str) and tic.strip() and isinstance(qty,(int,float)):
        moves[(typ,tic.strip())].append((d,qty))
    if isinstance(mv,str) and isinstance(g,(int,float)):
        if "ополн" in mv and d>"2024-01-01": flows[d]+=g
        elif "ывод" in mv: flows[d]-=abs(g)

# ---------- календарь и индекс MCFTR ----------
mcf=hist("stock","index","MCFTR","CLOSE")
cal=[d for d,_ in mcf if d>="2024-01-01"]
mcf_f=filler([(d,v["CLOSE"]) for d,v in mcf])

# ---------- цены ----------
print("Загрузка истории цен MOEX...")
stock_t=[t for (ty,t) in moves if ty=="Акции"]
bond_t=[t for (ty,t) in moves if ty=="Облигации"]
price={}     # ticker -> filler
bondmeta={}  # isin -> (facevalue_filler, faceunit)
SPLIT_TICKERS={"TRNFP","T"}   # известные сплиты: Транснефть 100:1 (2024), Т-Технологии 10:1 (2026). Учёт в Факте — в после-сплитных единицах.
for t in stock_t:
    h=hist("stock","shares",t,"CLOSE,LEGALCLOSEPRICE",board="TQBR")
    pr=[(d,(v["CLOSE"] if v["CLOSE"] is not None else v["LEGALCLOSEPRICE"])) for d,v in h]
    pr2=adjust_splits(pr) if t in SPLIT_TICKERS else sorted((d,v) for d,v in pr if v is not None)
    tag=" (сплит скорректирован)" if t in SPLIT_TICKERS else ""
    price[("Акции",t)]=filler(pr2); print(f"  акция {t}: {len(pr)} дней{tag}")
for isin in bond_t:
    h=hist("stock","bonds",isin,"CLOSE,LEGALCLOSEPRICE,FACEVALUE,FACEUNIT")
    # дедуп по дате: берём первую строку с ценой
    bydate={}
    for d,v in h:
        cl=v["CLOSE"] if v["CLOSE"] is not None else v["LEGALCLOSEPRICE"]
        if cl is not None and d not in bydate: bydate[d]=v|{"_px":cl}
    pr=[(d,v["_px"]) for d,v in bydate.items()]
    fv=[(d,v["FACEVALUE"]) for d,v in bydate.items() if v.get("FACEVALUE")]
    unit=None
    for d,v in sorted(bydate.items()):
        if v.get("FACEUNIT"): unit=v["FACEUNIT"]; break
    price[("Облигации",isin)]=filler(pr); bondmeta[isin]=(filler(fv) if fv else (lambda x:1000), unit or "SUR")
    print(f"  облиг {isin}: {len(pr)} дней, номинал-вал={unit}")
# LQDT
h=hist("stock","shares","LQDT","CLOSE,LEGALCLOSEPRICE",board="TQTF")
price[("Фонды","Фонд ликвидности LQDT")]=filler([(d,(v["CLOSE"] or v["LEGALCLOSEPRICE"])) for d,v in h]); print(f"  LQDT: {len(h)} дней")
# металлы + валюты
for sec in ["GLDRUB_TOM","SLVRUB_TOM"]:
    h=hist("currency","selt",sec,"CLOSE")
    price[("Другие активы",sec)]=filler([(d,v["CLOSE"]) for d,v in h]); print(f"  {sec}: {len(h)} дней")
fx={}
for unit,sec in [("USD","USD000UTSTOM"),("EUR","EUR_RUB__TOM"),("CNY","CNYRUB_TOM")]:
    h=hist("currency","selt",sec,"CLOSE")
    fx[unit]=filler([(d,v["CLOSE"]) for d,v in h]); print(f"  FX {unit}: {len(h)} дней")
fx["SUR"]=lambda d:1.0; fx["RUB"]=lambda d:1.0

# ---------- нетто-позиции по дням ----------
def netfunc(key):
    mv=sorted(moves[key]); ds_=[d for d,_ in mv]; cum=[]; s=0
    for _,q in mv: s+=q; cum.append(s)
    def f(date_str):
        i=bisect.bisect_right(ds_,date_str)-1
        return cum[i] if i>=0 else 0
    return f
netf={k:netfunc(k) for k in moves}

def value_on(date_str):
    tot=0
    for (typ,tic),nf in netf.items():
        q=nf(date_str)
        if abs(q)<1e-9: continue
        p=price.get((typ,tic))
        px=p(date_str) if p else None
        if px is None: continue
        if typ=="Облигации":
            fvf,unit=bondmeta[tic]; fv=fvf(date_str) or 1000
            tot+=q*(px/100.0)*fv*(fx.get(unit,fx["SUR"])(date_str) or 1)
        else:
            tot+=q*px
    return tot

daily=[(d,value_on(d)) for d in cal]
Vnow=daily[-1][1]

# ---------- дневной TWR ----------
def last_trading_on_or_before(target):
    i=bisect.bisect_right(cal,target)-1; return cal[i] if i>=0 else cal[0]
twr=[1.0]; rets=[]
for i in range(1,len(cal)):
    d=cal[i]; v0=daily[i-1][1]; v1=daily[i][1]; fl=flows.get(d,0)
    r=((v1-fl)/v0-1) if v0>0 else 0
    rets.append((d,r)); twr.append(twr[-1]*(1+r))
twr_f=dict(zip(cal,twr))
def cumret(date): return twr_f[last_trading_on_or_before(date)]-1
def period_twr(d0,d1):
    a=last_trading_on_or_before(d0); b=last_trading_on_or_before(d1)
    return twr_f[b]/twr_f[a]-1
end24=last_trading_on_or_before("2024-12-31"); end25=last_trading_on_or_before("2025-12-31")
r2024=period_twr("2024-01-01",end24); r2025=period_twr(end24,end25); r2026=period_twr(end25,cal[-1]); since=twr[-1]-1
# индекс MCFTR
def im_period(d0,d1):
    return mcf_f(last_trading_on_or_before(d1))/mcf_f(last_trading_on_or_before(d0))-1
im2024=im_period("2024-01-01",end24); im2025=im_period(end24,end25); im2026=im_period(end25,cal[-1])
imsince=mcf_f(cal[-1])/mcf_f(cal[0])-1

# ---------- линия (еженедельно), доходность % ----------
wk=cal[::5]
if wk[-1]!=cal[-1]: wk.append(cal[-1])
labels=[datetime.datetime.strptime(d,"%Y-%m-%d").strftime("%d.%m.%y") for d in wk]
port_line=[round(cumret(d)*100,1) for d in wk]
i0=mcf_f(cal[0]); index_line=[round((mcf_f(d)/i0-1)*100,1) for d in wk]

print(f"\n[диагностика TWR] V(start)={daily[0][1]:,.0f} V(end24)={value_on(end24):,.0f} V(end25)={value_on(end25):,.0f} V(now)={Vnow:,.0f}")
print(f"[диагностика TWR] 2024={r2024:.1%} 2025={r2025:.1%} 2026={r2026:.1%} с2024={since:.1%}")

# ============ ФИНАЛЬНАЯ СБОРКА ДАШБОРДА ============
it=av["Итоговая таблица"]; dk=av["Дивиденды и Купоны"]; sbsheet=av["Расчёт стоимости облигаций"]
NAMES={'ALRS':'АЛРОСА','LSNGP':'Россети Лен-п','MOEX':'МосБиржа','POSI':'Позитив','ROSN':'Роснефть',
 'RUAL':'РУСАЛ','SBER':'Сбербанк','SPBE':'СПБ Биржа','T':'Т-Технологии','TRNFP':'Транснефть ап','X5':'ИКС 5 (X5)'}
bondname={}
for r in range(2,60):
    a=sbsheet.cell(row=r,column=1).value; b=sbsheet.cell(row=r,column=2).value
    if isinstance(a,str): bondname[a]=b or a

# --- текущая оценка по категориям (последняя доступная цена) ---
today=cal[-1]
def val_asset(typ,tic,date):
    p=price.get((typ,tic)); px=p(date) if p else None; q=netf[(typ,tic)](date)
    if px is None or abs(q)<1e-9: return 0
    if typ=="Облигации":
        fvf,unit=bondmeta[tic]; fv=fvf(date) or 1000
        return q*(px/100.0)*fv*(fx.get(unit,fx["SUR"])(date) or 1)
    return q*px
cat=collections.defaultdict(float); stocks=[]; bonds=[]; lqdt_v=0; gold_v=0
for (typ,tic) in netf:
    v=val_asset(typ,tic,today)
    if v<=0: continue
    if typ=="Акции": cat["Акции"]+=v; stocks.append([NAMES.get(tic,tic),round(v)])
    elif typ=="Облигации": cat["Облигации"]+=v; bonds.append([bondname.get(tic,tic),round(v)])
    elif typ=="Фонды": cat["Фонды"]+=v; lqdt_v+=v
    elif typ=="Другие активы": cat["Золото"]+=v; gold_v+=v
stocks.sort(key=lambda x:-x[1]); bonds.sort(key=lambda x:-x[1])
cash_v=it["F2"].value or 0
total=cat["Акции"]+cat["Облигации"]+cat["Фонды"]+cat["Золото"]+cash_v

# --- доходность методом Дитца (логика с пополнениями), якоря — записанные значения ---
B11=it["B11"].value; B7=it["B7"].value; B44=it["B44"].value
contrib={2024:[],2025:[],2026:[]}
for r in range(2,38):
    y=it.cell(row=r,column=9).value; m=it.cell(row=r,column=10).value; k=it.cell(row=r,column=11).value
    if y in contrib and isinstance(m,(int,float)) and isinstance(k,(int,float)) and k: contrib[int(y)].append((int(m),k))
div2024=(dk["F15"].value or 0)+(dk["M12"].value or 0); div2025=(dk["F33"].value or 0)+(dk["L41"].value or 0); div2026=(dk["F60"].value or 0)+(dk["L60"].value or 0)
def mdietz(s,e,flows,divs,pm):
    nf=sum(a for _,a in flows); wc=sum(a*(pm-(m-1))/pm for m,a in flows if 1<=m<=pm)
    return (e-s-nf+divs)/(s+wc) if (s+wc) else 0
curm=datetime.date.today().month
# фантом LQDT, накопленный к концу 2025 (продажи ниже текущей цены завышали стоимость в таблице)
PXLQDT=2.0002; ph25=0.0
for r in range(3,fakt.max_row+1):
    tic=fakt.cell(row=r,column=3).value; q=fakt.cell(row=r,column=5).value; h=fakt.cell(row=r,column=8).value; dt=fakt.cell(row=r,column=1).value
    if isinstance(tic,str) and "LQDT" in tic and isinstance(q,(int,float)) and q<0 and isinstance(h,(int,float)) and hasattr(dt,'year') and dt<=datetime.datetime(2025,12,31):
        ph25+=(-q)*(PXLQDT-h)
# ГИБРИД: 2024/2025 — публичные цифры пользователя; 2026+ — корректно (исправленный LQDT)
mw24=mdietz(B11,B7,contrib[2024],div2024,12)
mw25=mdietz(B7,B44,contrib[2025],div2025,12)
mw26=mdietz(B44-ph25,total,contrib[2026],div2026,curm)   # старт2026 = конец2025 − фантом LQDT; конец = текущая исправл. стоимость
mwsince=(1+mw24)*(1+mw25)*(1+mw26)-1
print(f"[Дитц гибрид] 2024={mw24:.1%}(публ) 2025={mw25:.1%}(публ) 2026={mw26:.1%}(испр) ph25={ph25:,.0f} с2024={mwsince:.1%}")
print(f"[MCFTR]            2024={im2024:.1%} 2025={im2025:.1%} 2026={im2026:.1%} с2024={imsince:.1%}")

# --- диагностика: помесячная реконструированная стоимость по категориям ---
def breakdown(date):
    c=collections.defaultdict(float)
    for (typ,tic) in netf:
        v=val_asset(typ,tic,date)
        if v>0: c[typ]+=v
    return c
print("\n[диагностика помесячно] дата: total | акции | облигации | LQDT | золото")
for dm in ["2025-09-30","2025-10-31","2025-11-30","2025-12-30","2026-01-30","2026-02-27","2026-03-31","2026-04-15","2026-04-30","2026-05-15",cal[-1]]:
    dd=last_trading_on_or_before(dm); c=breakdown(dd)
    print(f"  {dd}: {sum(c.values()):>11,.0f} | {c['Акции']:>10,.0f} | {c['Облигации']:>10,.0f} | {c['Фонды']:>10,.0f} | {c['Другие активы']:>9,.0f}")
print("[поиск сплита] акции с резким падением цены за период:")
for (typ,tic) in sorted(netf):
    if typ!="Акции": continue
    p=price.get((typ,tic));
    if not p: continue
    for da,db in [("2026-04-10","2026-05-15"),("2026-01-20","2026-02-20"),("2025-06-01","2025-07-15")]:
        a=p(da); b=p(db)
        if a and b and a/b>3:
            print(f"  {tic}: {da} {a:.2f} -> {db} {b:.2f}  (x{a/b:.0f})  нетто_сейчас={netf[(typ,tic)](cal[-1]):.0f}")

# --- график: индекс MCFTR посуточно + портфель по опорным точкам доходности (надёжно) ---
import datetime as _dt
def od(s): return _dt.datetime.strptime(s,"%Y-%m-%d").toordinal()
acum=[(cal[0],0.0),(end24,mw24),(end25,(1+mw24)*(1+mw25)-1),(today,mwsince)]
aod=[(od(d),v) for d,v in acum]
def pcum(dd):
    o=od(dd)
    if o<=aod[0][0]: return aod[0][1]
    if o>=aod[-1][0]: return aod[-1][1]
    for i in range(1,len(aod)):
        if o<=aod[i][0]:
            (o0,v0),(o1,v1)=aod[i-1],aod[i]; return v0+(v1-v0)*(o-o0)/(o1-o0)
    return aod[-1][1]
wk=cal[::5]
if wk[-1]!=cal[-1]: wk.append(cal[-1])
labels=[_dt.datetime.strptime(d,"%Y-%m-%d").strftime("%d.%m.%y") for d in wk]
i0=mcf_f(cal[0]); index_line=[round((mcf_f(d)/i0-1)*100,1) for d in wk]
port_line=[round(pcum(d)*100,1) for d in wk]

# --- пополнения/выводы (из Факта, включая стартовый капитал) ---
contrib_list=[]
for r in range(3,fakt.max_row+1):
    dt=fakt.cell(row=r,column=1).value; mv=fakt.cell(row=r,column=4).value; g=fakt.cell(row=r,column=7).value
    if isinstance(mv,str) and isinstance(g,(int,float)) and hasattr(dt,'strftime'):
        if "ополн" in mv: contrib_list.append({"date":dt.strftime("%d.%m.%Y"),"type":"Пополнение","sum":round(g)})
        elif "ывод" in mv: contrib_list.append({"date":dt.strftime("%d.%m.%Y"),"type":"Вывод","sum":-round(abs(g))})

# --- история сделок + дивиденды/купоны (из журнала) ---
def fdate(v):
    if isinstance(v,_dt.datetime): return v.strftime("%d.%m.%Y")
    if v is None: return ""
    s=str(v).strip(); m=re.match(r'^(\d{1,2})\.(\d{1,2})\.(\d{2})$',s)
    return f"{int(m.group(1)):02d}.{int(m.group(2)):02d}.20{m.group(3)}" if m else s
namemap={'МАГНИТ':'Магнит','НОВАТЭК':'Новатэк','РУСАЛ':'Русал','АФК СИСТЕМА':'АФК Система','Polymetal':'Полиметалл',
 'TCSG':'Тинькофф','Тинькоф':'Тинькофф','Сбер':'Сбербанк','ГДР VK':'VK','Вконтакте':'VK','Группа Позитив':'Positive'}
wj=openpyxl.load_workbook(os.path.join(folder,"журнал долгосрочных сделок.xlsx"),data_only=True)
cfg=[("рынок РФ","Акции РФ",1,2,3,4,5,7,[(5,82,"Закрыта"),(85,102,"Открыта")]),
 ("рынок РФ","Облигации РФ",1,None,2,3,4,9,[(109,113,"Закрыта"),(117,124,"Открыта")]),
 ("зарубежные акции","Зарубежные",1,None,2,3,4,6,[(5,40,"Закрыта"),(43,47,"Открыта")]),
 ("криптовалюта","Крипта",1,2,3,4,5,9,[(5,21,"Закрыта")]),
 ("валюта","Валюта",1,None,2,3,4,6,[(5,10,"Закрыта"),(12,13,"Открыта")]),
 ("металлы","Металлы",1,2,3,4,5,9,[(5,6,"Закрыта"),(8,14,"Открыта")])]
trades=[]
for sh,cls,din,dout,ca,cp,cv,cr,ranges in cfg:
    ws=wj[sh]
    for r1,r2,st in ranges:
        for r in range(r1,r2+1):
            a=ws.cell(row=r,column=ca).value
            if not isinstance(a,str) or not a.strip(): continue
            a=namemap.get(a.strip(),a.strip())
            di=ws.cell(row=r,column=din).value; do=ws.cell(row=r,column=dout).value if dout else None
            if sh=="металлы" and r==6: di=_dt.datetime(2025,12,8); do=_dt.datetime(2026,1,6)
            pr=ws.cell(row=r,column=cp).value; vo=ws.cell(row=r,column=cv).value; rs=ws.cell(row=r,column=cr).value
            trades.append({"cls":cls,"asset":a,"din":fdate(di),"dout":fdate(do),
                "price":pr if isinstance(pr,(int,float)) else None,"vol":vo if isinstance(vo,(int,float)) else None,
                "res":round(rs,4) if isinstance(rs,(int,float)) else None,"status":st})
divs=[];coupons=[]
for r1,r2,yr in [(5,14,2024),(20,32,2025),(46,50,2026)]:
    for r in range(r1,r2+1):
        a=dk.cell(row=r,column=2).value; d=dk.cell(row=r,column=4).value; s=dk.cell(row=r,column=6).value
        if isinstance(a,str) and isinstance(s,(int,float)): divs.append({"year":yr,"asset":a,"date":fdate(d),"sum":round(s)})
for r1,r2,yr,cm in [(5,11,2024,13),(20,39,2025,12),(46,59,2026,12)]:
    for r in range(r1,r2+1):
        a=dk.cell(row=r,column=8).value; d=dk.cell(row=r,column=10).value; s=dk.cell(row=r,column=cm).value
        if isinstance(a,str) and isinstance(s,(int,float)): coupons.append({"year":yr,"asset":a,"date":fdate(d),"sum":round(s)})

DATA={
 "banner":"Котировки и индекс полной доходности (MCFTR) — Московская биржа, обновлено "+_dt.date.today().strftime("%d.%m.%Y")+". Состав и стоимость — из журнала движений. Доходность за 2024 и 2025 — как в исходном учёте; с 2026 — расчёт по уточнённой методике (с учётом пополнений). Сравнение — с индексом МосБиржи полной доходности.",
 "total":round(total),
 "classes":[["Акции",round(cat["Акции"]),"#60a5fa"],["Облигации",round(cat["Облигации"]),"#34d399"],
            ["Фонды (LQDT)",round(cat["Фонды"]),"#f59e0b"],["Золото",round(cat["Золото"]),"#a78bfa"],
            ["Свободные ДС",round(cash_v),"#64748b"]],
 "stocks":stocks,"bonds":bonds,
 "other":[["Фонд ликвидности LQDT",round(lqdt_v)],["Золото (GLDRUB_TOM)",round(gold_v)]],
 "kpi":{"value":round(total),"profit2024":round(total-B11-sum(a for fl in contrib.values() for _,a in fl)+div2024+div2025+div2026),
        "ytd2026":round(mw26,4),"y2025":round(mw25,4),"since2024":round(mwsince,4),"vsindex":round(mwsince-imsince,4)},
 "periods":[["2026 (с начала года)",round(mw26,4),round(im2026,4)],["2025 год",round(mw25,4),round(im2025,4)],
            ["2024 год",round(mw24,4),round(im2024,4)],["С 01.01.2024",round(mwsince,4),round(imsince,4)]],
 "line":{"labels":labels,"index":index_line,"portfolio":port_line},
 "contrib":contrib_list,"trades":trades,"divs":divs,"coupons":coupons,
}
json.dump(DATA,open(os.path.join(folder,"data.json"),"w",encoding="utf-8"),ensure_ascii=False)
try:
    tpl=open(os.path.join(folder,"_dashboard_template.html"),encoding="utf-8").read()
    html=tpl.replace("/*DATA*/","const DATA = "+json.dumps(DATA,ensure_ascii=False)+";")
    for fn in ("dashboard_prototype.html","index.html"):
        open(os.path.join(folder,fn),"w",encoding="utf-8").write(html)
except Exception as _e:
    print("HTML-сборка пропущена (для сервера не нужна):",_e)
print(f"\nИТОГО={round(total)} | акции={round(cat['Акции'])} обл={round(cat['Облигации'])} фонды={round(cat['Фонды'])} золото={round(cat['Золото'])}")
print(f"Пополнений/выводов: {len(contrib_list)} | сделок: {len(trades)}")
print("Дашборд собран: dashboard_prototype.html")
